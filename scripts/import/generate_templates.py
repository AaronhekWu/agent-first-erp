"""
ERP 数据导入模板生成器
生成 6 张 Excel 模板供各部门填写，填完后用 import_data.py 导入数据库

运行方式:
    pip install openpyxl
    python scripts/import/generate_templates.py

输出目录: scripts/import/templates/
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

OUTPUT_DIR = Path(__file__).parent / "templates"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── 颜色常量 ────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="C0392B")   # 必填列标题
BLUE_FILL  = PatternFill("solid", fgColor="2980B9")   # 可选列标题
GREY_FILL  = PatternFill("solid", fgColor="BDC3C7")   # 说明行背景
YELLOW_FILL= PatternFill("solid", fgColor="F9E79F")   # 示例行背景
GREEN_FILL = PatternFill("solid", fgColor="1ABC9C")   # Sheet 标签

WHITE_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
DARK_FONT  = Font(name="微软雅黑", size=9)
GREY_FONT  = Font(name="微软雅黑", size=9, color="7F8C8D", italic=True)
BOLD_FONT  = Font(name="微软雅黑", bold=True, size=9)

THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header(cell, required: bool):
    cell.fill  = RED_FILL if required else BLUE_FILL
    cell.font  = WHITE_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_note(cell):
    cell.fill  = GREY_FILL
    cell.font  = GREY_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def style_example(cell):
    cell.fill  = YELLOW_FILL
    cell.font  = DARK_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def style_data(cell):
    cell.font  = DARK_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def add_dropdown(ws, col_letter, start_row, end_row, formula: str, prompt: str):
    dv = DataValidation(
        type="list", formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="输入错误",
        error=f"请从下拉列表选择: {prompt}",
        showInputMessage=True,
        promptTitle="可选值",
        prompt=prompt,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


def freeze_and_filter(ws, freeze_at="A4"):
    ws.freeze_panes = freeze_at
    ws.auto_filter.ref = ws.dimensions


def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_legend(wb):
    ws = wb.create_sheet("图例说明")
    ws.sheet_properties.tabColor = "95A5A6"
    items = [
        ("红色标题", "必填字段，不可为空"),
        ("蓝色标题", "可选字段，留空使用默认值"),
        ("灰色行",   "字段说明，填写时请删除本行"),
        ("黄色行",   "示例数据，填写时请删除本行"),
        ("白色行",   "实际填写区域"),
    ]
    ws["A1"] = "图例说明"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=12)
    for i, (name, desc) in enumerate(items, 3):
        ws.cell(i, 1, name).font = BOLD_FONT
        ws.cell(i, 2, desc).font = DARK_FONT
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40


# ══════════════════════════════════════════════════════════════
# T1: 学员信息（含家长）
# ══════════════════════════════════════════════════════════════
def make_t1_students(wb):
    ws = wb.create_sheet("T1_学员信息")
    ws.sheet_properties.tabColor = "E74C3C"

    COLUMNS = [
        # (标题, 字段说明, 示例值, required, col_width)
        ("姓名*",         "学员真实姓名，不超过50字",           "张小明",       True,  14),
        ("手机号",        "学员本人手机号（11位）",              "13800138001",  False, 14),
        ("性别",          "填: 男 或 女",                        "男",           False, 8),
        ("出生日期",      "格式: 2012-05-20",                    "2012-05-20",   False, 14),
        ("学校",          "就读学校名称",                        "北京第一中学", False, 20),
        ("年级",          "小学/初一/初二/初三/高一/高二/高三",  "初一",         False, 10),
        ("来源渠道",      "微信/转介绍/自然到访/电话/其他",      "微信",         False, 12),
        ("负责顾问姓名",  "填写系统中顾问的姓名（需已建账号）",  "李老师",       False, 16),
        ("所属部门",      "管理部/市场部/教学部/财务部",         "教学部",       False, 12),
        ("状态",          "在读/停课/毕业（默认：在读）",        "在读",         False, 10),
        ("备注",          "其他补充信息",                        "对数学有兴趣", False, 24),
        # 家长信息
        ("家长姓名",      "主要联系人姓名",                      "张父",         False, 12),
        ("家长手机",      "家长联系电话",                        "13800138002",  False, 14),
        ("家长关系",      "父亲/母亲/爷爷/奶奶/其他",            "父亲",         False, 10),
        ("家长微信号",    "家长微信（选填）",                    "zhang_dad",    False, 14),
    ]

    # 行1：表头
    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    # 行2：说明
    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        cell = ws.cell(2, c, f"[说明] {note}")
        style_note(cell)
    ws.row_dimensions[2].height = 20

    # 行3：示例
    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        cell = ws.cell(3, c, example)
        style_example(cell)
    ws.row_dimensions[3].height = 20

    # 行4-1003：数据区
    for row in range(4, 1004):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    # 下拉验证
    add_dropdown(ws, "C", 4, 1003, '"男,女"',              "男/女")
    add_dropdown(ws, "F", 4, 1003,
                 '"小学,初一,初二,初三,高一,高二,高三"',    "年级")
    add_dropdown(ws, "G", 4, 1003,
                 '"微信,转介绍,自然到访,电话,其他"',        "来源渠道")
    add_dropdown(ws, "I", 4, 1003,
                 '"管理部,市场部,教学部,财务部"',            "部门")
    add_dropdown(ws, "J", 4, 1003, '"在读,停课,毕业"',      "状态")
    add_dropdown(ws, "N", 4, 1003,
                 '"父亲,母亲,爷爷,奶奶,其他"',              "关系")

    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════════════
# T2: 课程信息
# ══════════════════════════════════════════════════════════════
def make_t2_courses(wb):
    ws = wb.create_sheet("T2_课程信息")
    ws.sheet_properties.tabColor = "E67E22"

    COLUMNS = [
        ("课程名称*",     "完整课程名称",                       "2026春季数学培训班", True,  22),
        ("学科",          "数学/语文/英语/物理/化学/其他",       "数学",              False, 12),
        ("适用年级",      "小学/初一/初二/初三/高一/高二/高三",  "初一",              False, 12),
        ("班级容量",      "最大学员人数（整数），空=不限",        "20",                False, 10),
        ("单课费用",      "每节课费用（元），如 200.00",         "200.00",            False, 12),
        ("开始日期",      "格式: 2026-05-01",                   "2026-05-01",        False, 14),
        ("结束日期",      "格式: 2026-07-31",                   "2026-07-31",        False, 14),
        ("上课星期",      "1=周一 … 7=周日",                    "3",                 False, 10),
        ("上课时间",      "格式: 15:00",                        "15:00",             False, 10),
        ("课程时长(分钟)","每节课分钟数",                        "90",                False, 14),
        ("所属部门",      "管理部/市场部/教学部/财务部",         "教学部",            False, 12),
        ("状态",          "开放/暂停/归档（默认：开放）",        "开放",              False, 10),
        ("描述",          "课程简介",                           "专注初一数学…",     False, 30),
    ]

    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        style_note(ws.cell(2, c, f"[说明] {note}"))

    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        style_example(ws.cell(3, c, example))

    for row in range(4, 504):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    add_dropdown(ws, "B", 4, 503,
                 '"数学,语文,英语,物理,化学,历史,地理,生物,其他"', "学科")
    add_dropdown(ws, "C", 4, 503,
                 '"小学,初一,初二,初三,高一,高二,高三"', "年级")
    add_dropdown(ws, "K", 4, 503,
                 '"管理部,市场部,教学部,财务部"', "部门")
    add_dropdown(ws, "L", 4, 503, '"开放,暂停,归档"', "状态")

    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════════════
# T3: 报名记录
# ══════════════════════════════════════════════════════════════
def make_t3_enrollments(wb):
    ws = wb.create_sheet("T3_报名记录")
    ws.sheet_properties.tabColor = "27AE60"

    COLUMNS = [
        ("学员姓名*",     "须与T1学员信息中的姓名完全一致",    "张小明",       True,  16),
        ("学员手机号",    "如有重名，用手机号区分",             "13800138001",  False, 14),
        ("课程名称*",     "须与T2课程信息中的名称完全一致",    "2026春季数学培训班", True, 22),
        ("报名状态",      "在读/已完成/已取消（默认：在读）",   "在读",         False, 12),
        ("单价(元)",      "每节课价格，空=取课程默认费用",       "200.00",       False, 12),
        ("总课时数",      "本次报名总课时",                    "30",           False, 10),
        ("已消课时",      "截止导入已上课时数",                "5",            False, 10),
        ("总金额(元)",    "本次报名总费用",                    "6000.00",      False, 12),
        ("报名日期",      "格式: 2026-05-01（空=导入当天）",   "2026-05-01",   False, 14),
        ("来源",          "正常/转介绍/转班",                  "正常",         False, 10),
        ("备注",          "备注信息",                          "",             False, 20),
    ]

    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        style_note(ws.cell(2, c, f"[说明] {note}"))

    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        style_example(ws.cell(3, c, example))

    for row in range(4, 1004):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    add_dropdown(ws, "D", 4, 1003, '"在读,已完成,已取消"', "状态")
    add_dropdown(ws, "J", 4, 1003, '"正常,转介绍,转班"',   "来源")

    freeze_and_filter(ws, "A4")

    # 注意说明
    ws["A1005"] = "⚠ 注意：学员姓名与课程名称须与对应模板中完全一致，系统将按名称+手机号匹配。重名学员必须填写手机号。"
    ws["A1005"].font = Font(name="微软雅黑", color="C0392B", bold=True, size=9)


# ══════════════════════════════════════════════════════════════
# T4: 充值记录（历史数据）
# ══════════════════════════════════════════════════════════════
def make_t4_recharges(wb):
    ws = wb.create_sheet("T4_充值记录")
    ws.sheet_properties.tabColor = "8E44AD"

    COLUMNS = [
        ("学员姓名*",     "须与T1完全一致",                    "张小明",       True,  16),
        ("学员手机号",    "重名学员必填",                       "13800138001",  False, 14),
        ("充值金额*",     "实收金额（元），如 2000.00",         "2000.00",      True,  14),
        ("赠送金额",      "赠送/赠课折算金额，无则留空",        "200.00",       False, 12),
        ("支付方式*",     "现金/微信/支付宝/银行转账/其他",    "微信",         True,  12),
        ("支付流水号",    "第三方支付单号（选填）",             "WX202605010001", False, 20),
        ("充值日期",      "格式: 2026-05-01（空=导入当天）",   "2026-05-01",   False, 14),
        ("备注",          "充值说明",                          "春季促销充值", False, 24),
    ]

    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        style_note(ws.cell(2, c, f"[说明] {note}"))

    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        style_example(ws.cell(3, c, example))

    for row in range(4, 1004):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    add_dropdown(ws, "E", 4, 1003,
                 '"现金,微信,支付宝,银行转账,其他"', "支付方式")

    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════════════
# T5: 跟进记录
# ══════════════════════════════════════════════════════════════
def make_t5_followups(wb):
    ws = wb.create_sheet("T5_跟进记录")
    ws.sheet_properties.tabColor = "16A085"

    COLUMNS = [
        ("学员姓名*",     "须与T1完全一致",                         "张小明",          True,  16),
        ("学员手机号",    "重名学员必填",                            "13800138001",     False, 14),
        ("跟进方式*",     "电话/微信/上门/其他",                    "微信",            True,  10),
        ("跟进内容*",     "本次跟进的详细描述（至少5字）",           "和家长沟通续费意向，家长表示感兴趣", True, 40),
        ("跟进结果",      "跟进结果描述",                           "有续费意向",      False, 20),
        ("下次计划",      "下次跟进计划",                           "下周电话回访确认", False, 24),
        ("下次跟进日期",  "格式: 2026-05-10",                       "2026-05-10",      False, 16),
        ("跟进日期",      "格式: 2026-04-21（空=导入当天）",        "2026-04-21",      False, 14),
        ("跟进人姓名",    "填写顾问姓名（需已建账号），空=导入人",  "李老师",          False, 16),
    ]

    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        style_note(ws.cell(2, c, f"[说明] {note}"))

    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        style_example(ws.cell(3, c, example))

    for row in range(4, 2004):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    add_dropdown(ws, "C", 4, 2003, '"电话,微信,上门,其他"', "跟进方式")

    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════════════
# T6: 考勤记录（历史补录）
# ══════════════════════════════════════════════════════════════
def make_t6_attendance(wb):
    ws = wb.create_sheet("T6_考勤记录")
    ws.sheet_properties.tabColor = "2C3E50"

    COLUMNS = [
        ("学员姓名*",     "须与T1完全一致",                    "张小明",       True,  16),
        ("学员手机号",    "重名学员必填",                       "13800138001",  False, 14),
        ("课程名称*",     "须与T2完全一致",                    "2026春季数学培训班", True, 22),
        ("上课日期*",     "格式: 2026-04-21",                  "2026-04-21",   True,  14),
        ("考勤状态*",     "出勤/缺勤/迟到/请假",               "出勤",         True,  10),
        ("是否消课",      "是/否（出勤时默认：是）",            "是",           False, 10),
        ("备注",          "请假原因等",                        "",             False, 24),
    ]

    for c, (title, _, __, req, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        style_header(cell, req)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 32

    for c, (_, note, __, ___, ____) in enumerate(COLUMNS, 1):
        style_note(ws.cell(2, c, f"[说明] {note}"))

    for c, (_, __, example, ___, ____) in enumerate(COLUMNS, 1):
        style_example(ws.cell(3, c, example))

    for row in range(4, 2004):
        for col in range(1, len(COLUMNS) + 1):
            style_data(ws.cell(row, col))

    add_dropdown(ws, "E", 4, 2003, '"出勤,缺勤,迟到,请假"', "考勤状态")
    add_dropdown(ws, "F", 4, 2003, '"是,否"',               "是否消课")

    freeze_and_filter(ws, "A4")

    ws["A2005"] = "⚠ 说明：考勤记录一般在系统上线后实时录入，此模板仅用于历史数据补录。"
    ws["A2005"].font = Font(name="微软雅黑", color="7F8C8D", italic=True, size=9)


# ══════════════════════════════════════════════════════════════
# 主入口
# ══════════════════════════════════════════════════════════════
def generate_all():
    print("开始生成模板...")

    templates = [
        ("T1_学员信息",   make_t1_students),
        ("T2_课程信息",   make_t2_courses),
        ("T3_报名记录",   make_t3_enrollments),
        ("T4_充值记录",   make_t4_recharges),
        ("T5_跟进记录",   make_t5_followups),
        ("T6_考勤记录",   make_t6_attendance),
    ]

    for name, func in templates:
        wb = Workbook()
        wb.remove(wb.active)    # 删除默认 Sheet
        func(wb)
        add_legend(wb)
        path = OUTPUT_DIR / f"{name}.xlsx"
        wb.save(path)
        print(f"  [OK] {path}")

    # 生成汇总包（所有模板在一个 workbook）
    wb_all = Workbook()
    wb_all.remove(wb_all.active)
    for _, func in templates:
        func(wb_all)
    add_legend(wb_all)
    all_path = OUTPUT_DIR / "ERP数据导入_全部模板合集.xlsx"
    wb_all.save(all_path)
    print(f"  [OK] {all_path}")

    print(f"\n完成！共生成 {len(templates) + 1} 个文件，保存在: {OUTPUT_DIR}")
    print("\n填写顺序建议:")
    print("  1. T1_学员信息  → 2. T2_课程信息  → 3. T3_报名记录")
    print("  4. T4_充值记录  → 5. T5_跟进记录  → 6. T6_考勤记录（可选）")


if __name__ == "__main__":
    generate_all()
