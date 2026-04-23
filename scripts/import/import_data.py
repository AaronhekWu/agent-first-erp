"""
ERP 数据导入脚本
读取填好的 Excel 模板，调用 Supabase RPC 导入数据库

运行方式:
    pip install openpyxl supabase python-dotenv
    python scripts/import/import_data.py --sheet T1 --file templates/T1_学员信息.xlsx
    python scripts/import/import_data.py --all   # 按顺序导入全部模板

导入顺序: T1 → T2 → T3 → T4 → T5 → T6

选项:
    --sheet T1|T2|T3|T4|T5|T6  单独导入某张表
    --file  <path>              指定 Excel 文件路径
    --all                       按顺序导入 templates/ 目录下全部模板
    --dry-run                   仅校验不写库
    --skip-errors               遇到行错误继续（不中止）
    --batch-size 50             每批并发数（默认 50）
"""

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装依赖: pip install openpyxl supabase python-dotenv")
    sys.exit(1)

try:
    from supabase import create_client, Client
    from dotenv import load_dotenv
except ImportError:
    print("请先安装依赖: pip install supabase python-dotenv")
    sys.exit(1)

import os

# ── 配置 ──────────────────────────────────────────────────────
load_dotenv(Path(__file__).parents[2] / ".env.local")
load_dotenv(Path(__file__).parents[2] / ".env")

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "http://47.102.28.236:80")
# 导入脚本使用 SERVICE_ROLE KEY（绕过 RLS），需在 .env.local 中配置
SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

TEMPLATES_DIR = Path(__file__).parent / "templates"

# ── 枚举映射（中文 → 英文）────────────────────────────────────
GENDER_MAP = {"男": "male", "女": "female"}
STATUS_MAP = {"在读": "active", "停课": "inactive", "毕业": "graduated"}
SOURCE_MAP = {"微信": "wechat", "转介绍": "referral", "自然到访": "walk_in",
              "电话": "phone", "其他": "other"}
PAYMENT_MAP = {"现金": "cash", "微信": "wechat", "支付宝": "alipay",
               "银行转账": "bank_transfer", "其他": "other"}
FOLLOWUP_TYPE_MAP = {"电话": "phone", "微信": "wechat", "上门": "visit", "其他": "other"}
ATTEND_STATUS_MAP = {"出勤": "present", "缺勤": "absent", "迟到": "late", "请假": "leave"}
ENROLL_STATUS_MAP = {"在读": "enrolled", "已完成": "completed",
                     "已取消": "cancelled", "转班": "transferred"}
ENROLL_SOURCE_MAP = {"正常": "normal", "转介绍": "referral", "转班": "transfer"}
DEPT_MAP = {"管理部": "8e0422de-f874-468e-8efa-5573810adbfc",
            "市场部": "07f828ff-4aaa-49d9-b453-d97c490eacc3",
            "教学部": "edb7a501-ddf9-4649-b44f-cd04cf9bd111",
            "财务部": "0a731438-123a-45ea-a532-89426f1a456e"}
COURSE_STATUS_MAP = {"开放": "active", "暂停": "inactive", "归档": "archived"}

# ── 工具函数 ──────────────────────────────────────────────────

def clean(v) -> str | None:
    """清理单元格值"""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_date(v) -> str | None:
    """解析日期，返回 YYYY-MM-DD"""
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    # 支持 2026/05/01 或 2026-05-01 或 20260501
    s = re.sub(r"[/\.]", "-", s)
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except ValueError:
        raise ValueError(f"日期格式不正确: {v!r}，请使用 YYYY-MM-DD")


def parse_decimal(v, field: str) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).strip())
    except ValueError:
        raise ValueError(f"{field} 必须是数字，得到: {v!r}")


def parse_int(v, field: str) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(str(v).strip())
    except ValueError:
        raise ValueError(f"{field} 必须是整数，得到: {v!r}")


def map_enum(v, mapping: dict, field: str, required=False) -> str | None:
    if v is None or str(v).strip() == "":
        if required:
            raise ValueError(f"{field} 不能为空")
        return None
    s = str(v).strip()
    if s not in mapping:
        raise ValueError(f"{field} 值 '{s}' 不合法，可选: {list(mapping.keys())}")
    return mapping[s]


def read_sheet(path: Path, sheet_name_hint: str = None):
    """读取 Excel，跳过说明行(2)和示例行(3)，从第4行开始"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name_hint:
        ws = next((wb[s] for s in wb.sheetnames if sheet_name_hint in s), None)
        if not ws:
            raise ValueError(f"找不到包含 '{sheet_name_hint}' 的工作表，现有: {wb.sheetnames}")
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return []  # 无数据

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[3:]  # 跳过说明行和示例行

    result = []
    for i, row in enumerate(data_rows, start=4):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # 跳过空行
        result.append({"_row": i, **dict(zip(headers, row))})

    return result


# ── 学员缓存（name+phone → id）────────────────────────────────
_student_cache: dict[str, str] = {}
_counselor_cache: dict[str, str] = {}
_course_cache: dict[str, str] = {}


def get_student_id(supabase: Client, name: str, phone: str | None = None) -> str:
    key = f"{name}|{phone or ''}"
    if key in _student_cache:
        return _student_cache[key]

    q = supabase.from_("stu_students").select("id").eq("name", name).is_("deleted_at", "null")
    if phone:
        q = q.eq("phone", phone)
    res = q.execute()
    if not res.data:
        raise ValueError(f"找不到学员: '{name}' (手机: {phone})，请先导入 T1_学员信息")
    if len(res.data) > 1:
        raise ValueError(f"学员 '{name}' 有 {len(res.data)} 个匹配，请补填手机号区分")
    sid = res.data[0]["id"]
    _student_cache[key] = sid
    return sid


def get_counselor_id(supabase: Client, name: str) -> str | None:
    if not name:
        return None
    if name in _counselor_cache:
        return _counselor_cache[name]
    res = supabase.from_("acct_profiles").select("id").eq("display_name", name).execute()
    if not res.data:
        raise ValueError(f"找不到顾问: '{name}'，请确认已在系统建账号")
    cid = res.data[0]["id"]
    _counselor_cache[name] = cid
    return cid


def get_course_id(supabase: Client, name: str) -> str:
    if name in _course_cache:
        return _course_cache[name]
    res = supabase.from_("crs_courses").select("id").eq("name", name).is_("deleted_at", "null").execute()
    if not res.data:
        raise ValueError(f"找不到课程: '{name}'，请先导入 T2_课程信息")
    if len(res.data) > 1:
        raise ValueError(f"课程名称 '{name}' 有多个匹配，请确认课程名唯一")
    cid = res.data[0]["id"]
    _course_cache[name] = cid
    return cid


def get_enrollment_id(supabase: Client, student_id: str, course_id: str) -> str:
    res = (supabase.from_("crs_enrollments")
           .select("id")
           .eq("student_id", student_id)
           .eq("course_id", course_id)
           .neq("status", "cancelled")
           .execute())
    if not res.data:
        raise ValueError("找不到报名记录，请先导入 T3_报名记录")
    return res.data[0]["id"]


# ── 各模板导入函数 ─────────────────────────────────────────────

def import_t1(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T1 学员信息"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name = clean(r.get("姓名*"))
            if not name:
                raise ValueError("姓名不能为空")

            params = {
                "p_name":          name,
                "p_phone":         clean(r.get("手机号")),
                "p_gender":        map_enum(r.get("性别"), GENDER_MAP, "性别"),
                "p_birth_date":    parse_date(r.get("出生日期")),
                "p_school":        clean(r.get("学校")),
                "p_grade":         clean(r.get("年级")),
                "p_source":        map_enum(r.get("来源渠道"), SOURCE_MAP, "来源渠道"),
                "p_status":        map_enum(r.get("状态"), STATUS_MAP, "状态") or "active",
                "p_notes":         clean(r.get("备注")),
                "p_department_id": DEPT_MAP.get(str(r.get("所属部门", "")).strip()),
                "p_parent_name":   clean(r.get("家长姓名")),
                "p_parent_phone":  clean(r.get("家长手机")),
                "p_parent_relation": clean(r.get("家长关系")),
                "p_operator_id":   operator_id,
            }

            counselor_name = clean(r.get("负责顾问姓名"))
            if counselor_name:
                params["p_assigned_to"] = get_counselor_id(supabase, counselor_name)

            print(f"  行 {row_num}: 学员 [{name}]", end="")
            if not dry_run:
                res = supabase.rpc("rpc_create_student", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))
                # 缓存新建学员
                sid = res.data.get("student_id") if res.data else None
                if sid:
                    _student_cache[f"{name}|{params.get('p_phone', '')}"] = sid
            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


def import_t2(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T2 课程信息"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name = clean(r.get("课程名称*"))
            if not name:
                raise ValueError("课程名称不能为空")

            weekday = parse_int(r.get("上课星期"), "上课星期")
            time_   = clean(r.get("上课时间"))
            dur     = parse_int(r.get("课程时长(分钟)"), "课程时长")
            schedule = {}
            if weekday is not None: schedule["weekday"] = weekday
            if time_:               schedule["time"] = time_
            if dur is not None:     schedule["duration_minutes"] = dur

            params = {
                "p_name":          name,
                "p_subject":       clean(r.get("学科")),
                "p_level":         clean(r.get("适用年级")),
                "p_max_capacity":  parse_int(r.get("班级容量"), "班级容量"),
                "p_fee":           parse_decimal(r.get("单课费用"), "单课费用"),
                "p_start_date":    parse_date(r.get("开始日期")),
                "p_end_date":      parse_date(r.get("结束日期")),
                "p_schedule_info": schedule or None,
                "p_department_id": DEPT_MAP.get(str(r.get("所属部门", "")).strip()),
                "p_description":   clean(r.get("描述")),
                "p_operator_id":   operator_id,
            }

            print(f"  行 {row_num}: 课程 [{name}]", end="")
            if not dry_run:
                res = supabase.rpc("rpc_create_course", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))
            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


def import_t3(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T3 报名记录"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name   = clean(r.get("学员姓名*"))
            course = clean(r.get("课程名称*"))
            if not name or not course:
                raise ValueError("学员姓名和课程名称不能为空")

            phone = clean(r.get("学员手机号"))
            student_id = get_student_id(supabase, name, phone)
            course_id  = get_course_id(supabase, course)

            params = {
                "p_student_id": student_id,
                "p_course_id":  course_id,
                "p_source":     ENROLL_SOURCE_MAP.get(
                    str(r.get("来源", "")).strip(), "normal"),
                "p_notes":      clean(r.get("备注")),
                "p_operator_id": operator_id,
            }

            print(f"  行 {row_num}: [{name}] → [{course}]", end="")
            if not dry_run:
                res = supabase.rpc("rpc_enroll_student", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))

                # 如有历史消课数据，直接更新 enrollment 字段
                consumed = parse_int(r.get("已消课时"), "已消课时")
                total    = parse_int(r.get("总课时数"), "总课时数")
                unit_p   = parse_decimal(r.get("单价(元)"), "单价")
                total_a  = parse_decimal(r.get("总金额(元)"), "总金额")
                enroll_id = res.data.get("enrollment_id") if res.data else None

                update_data: dict[str, Any] = {}
                if consumed is not None: update_data["consumed_lessons"] = consumed
                if total is not None:    update_data["total_lessons"] = total
                if unit_p is not None:   update_data["unit_price"] = unit_p
                if total_a is not None:  update_data["total_amount"] = total_a
                if update_data and enroll_id:
                    supabase.from_("crs_enrollments").update(update_data).eq("id", enroll_id).execute()

                status = map_enum(r.get("报名状态"), ENROLL_STATUS_MAP, "报名状态")
                if status and status != "enrolled" and enroll_id:
                    supabase.from_("crs_enrollments").update({"status": status}).eq("id", enroll_id).execute()

            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


def import_t4(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T4 充值记录"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name   = clean(r.get("学员姓名*"))
            amount = parse_decimal(r.get("充值金额*"), "充值金额")
            method = map_enum(r.get("支付方式*"), PAYMENT_MAP, "支付方式", required=True)
            if not name or amount is None:
                raise ValueError("学员姓名和充值金额不能为空")

            phone = clean(r.get("学员手机号"))
            student_id = get_student_id(supabase, name, phone)

            params = {
                "p_student_id":     student_id,
                "p_amount":         amount,
                "p_payment_method": method,
                "p_bonus_amount":   parse_decimal(r.get("赠送金额"), "赠送金额") or 0.0,
                "p_payment_ref":    clean(r.get("支付流水号")),
                "p_notes":          clean(r.get("备注")),
                "p_operator_id":    operator_id,
            }

            print(f"  行 {row_num}: [{name}] 充值 {amount}元 ({method})", end="")
            if not dry_run:
                res = supabase.rpc("rpc_recharge", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))
            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


def import_t5(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T5 跟进记录"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name    = clean(r.get("学员姓名*"))
            ftype   = map_enum(r.get("跟进方式*"), FOLLOWUP_TYPE_MAP, "跟进方式", required=True)
            content = clean(r.get("跟进内容*"))
            if not name or not content:
                raise ValueError("学员姓名和跟进内容不能为空")
            if len(content) < 5:
                raise ValueError("跟进内容至少5个字")

            phone = clean(r.get("学员手机号"))
            student_id = get_student_id(supabase, name, phone)

            counselor_id = operator_id
            counselor_name = clean(r.get("跟进人姓名"))
            if counselor_name:
                counselor_id = get_counselor_id(supabase, counselor_name) or operator_id

            next_date = parse_date(r.get("下次跟进日期"))
            params = {
                "p_student_id":  student_id,
                "p_type":        ftype,
                "p_content":     content,
                "p_result":      clean(r.get("跟进结果")),
                "p_next_plan":   clean(r.get("下次计划")),
                "p_next_date":   f"{next_date}T09:00:00+08:00" if next_date else None,
                "p_operator_id": counselor_id,
            }

            print(f"  行 {row_num}: [{name}] {ftype}", end="")
            if not dry_run:
                res = supabase.rpc("rpc_create_followup", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))
            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


def import_t6(supabase: Client, rows: list, dry_run: bool, skip_errors: bool, operator_id: str):
    """T6 考勤记录"""
    ok = err = 0
    for r in rows:
        row_num = r["_row"]
        try:
            name   = clean(r.get("学员姓名*"))
            course = clean(r.get("课程名称*"))
            cdate  = parse_date(r.get("上课日期*"))
            status = map_enum(r.get("考勤状态*"), ATTEND_STATUS_MAP, "考勤状态", required=True)
            if not name or not course or not cdate:
                raise ValueError("学员姓名、课程名称、上课日期不能为空")

            phone = clean(r.get("学员手机号"))
            student_id = get_student_id(supabase, name, phone)
            course_id  = get_course_id(supabase, course)
            enrollment_id = get_enrollment_id(supabase, student_id, course_id)

            trigger_consume = str(r.get("是否消课", "是")).strip() == "是" and status == "present"

            params = {
                "p_enrollment_id":   enrollment_id,
                "p_class_date":      cdate,
                "p_status":          status,
                "p_trigger_consume": trigger_consume,
                "p_notes":           clean(r.get("备注")),
                "p_operator_id":     operator_id,
            }

            print(f"  行 {row_num}: [{name}] {course} {cdate} {status}", end="")
            if not dry_run:
                res = supabase.rpc("rpc_mark_attendance", params).execute()
                if hasattr(res, "error") and res.error:
                    raise ValueError(str(res.error))
            print(" ✓")
            ok += 1
        except Exception as e:
            print(f" ✗ {e}")
            err += 1
            if not skip_errors:
                break
    return ok, err


SHEET_MAP = {
    "T1": ("T1_学员信息",   import_t1),
    "T2": ("T2_课程信息",   import_t2),
    "T3": ("T3_报名记录",   import_t3),
    "T4": ("T4_充值记录",   import_t4),
    "T5": ("T5_跟进记录",   import_t5),
    "T6": ("T6_考勤记录",   import_t6),
}

DEFAULT_FILES = {
    "T1": "T1_学员信息.xlsx",
    "T2": "T2_课程信息.xlsx",
    "T3": "T3_报名记录.xlsx",
    "T4": "T4_充值记录.xlsx",
    "T5": "T5_跟进记录.xlsx",
    "T6": "T6_考勤记录.xlsx",
}


def run_import(sheet_key: str, file_path: Path, supabase: Client,
               dry_run: bool, skip_errors: bool, operator_id: str):
    hint, func = SHEET_MAP[sheet_key]
    print(f"\n{'[DRY-RUN] ' if dry_run else ''}导入 {hint}（{file_path.name}）")
    print("─" * 60)

    rows = read_sheet(file_path, hint)
    if not rows:
        print("  ⚠ 无数据行，跳过")
        return 0, 0

    print(f"  共 {len(rows)} 行数据，开始导入...\n")
    ok, err = func(supabase, rows, dry_run, skip_errors, operator_id)
    print(f"\n  结果: {ok} 成功 / {err} 失败")
    return ok, err


def main():
    parser = argparse.ArgumentParser(description="ERP 数据导入工具")
    parser.add_argument("--sheet",       choices=["T1","T2","T3","T4","T5","T6"], help="单独导入指定模板")
    parser.add_argument("--file",        help="指定 Excel 文件路径")
    parser.add_argument("--all",         action="store_true", help="按顺序导入全部模板")
    parser.add_argument("--dry-run",     action="store_true", help="仅校验不写库")
    parser.add_argument("--skip-errors", action="store_true", help="遇错跳过继续")
    parser.add_argument("--operator-id", default=None,
                        help="操作人 UUID（acct_profiles.id），不填则使用 service_role 身份")
    args = parser.parse_args()

    if not args.sheet and not args.all:
        parser.print_help()
        sys.exit(1)

    if not SERVICE_KEY:
        print("错误: 未设置 SUPABASE_SERVICE_ROLE_KEY 环境变量")
        print("请在 .env.local 中添加: SUPABASE_SERVICE_ROLE_KEY=your_service_role_key")
        sys.exit(1)

    supabase: Client = create_client(SUPABASE_URL, SERVICE_KEY)

    # 使用第一个 admin 作为默认 operator
    operator_id = args.operator_id
    if not operator_id:
        res = supabase.from_("acct_profiles").select("id").eq("is_active", True).limit(1).execute()
        if res.data:
            operator_id = res.data[0]["id"]
        else:
            print("错误: 系统中无可用用户，请先在系统中创建管理员账号")
            sys.exit(1)
    print(f"操作人 ID: {operator_id}")

    total_ok = total_err = 0

    if args.all:
        for key in ["T1", "T2", "T3", "T4", "T5", "T6"]:
            fp = TEMPLATES_DIR / DEFAULT_FILES[key]
            if not fp.exists():
                print(f"\n⚠ 跳过 {key}：文件不存在 ({fp})")
                continue
            ok, err = run_import(key, fp, supabase, args.dry_run, args.skip_errors, operator_id)
            total_ok += ok
            total_err += err
    else:
        fp = Path(args.file) if args.file else TEMPLATES_DIR / DEFAULT_FILES[args.sheet]
        if not fp.exists():
            print(f"错误: 文件不存在: {fp}")
            sys.exit(1)
        ok, err = run_import(args.sheet, fp, supabase, args.dry_run, args.skip_errors, operator_id)
        total_ok, total_err = ok, err

    print("\n" + "═" * 60)
    print(f"导入完成: 总成功 {total_ok} 行 / 总失败 {total_err} 行")
    if total_err > 0:
        print("请检查以上 ✗ 行并修正后重新导入（已成功的行不会重复插入）")
    sys.exit(1 if total_err > 0 else 0)


if __name__ == "__main__":
    main()
